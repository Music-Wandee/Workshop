from gensim.models import Word2Vec
from openpyxl import Workbook

def write_ver_excel(model):
    # write report (Vertical)
    vertical_similarity = list()
    words = list(model.wv.vocab)
    for i in range(len(words)):
        for j in range(len(words)):
            vec = model.wv.similarity(words[i],words[j])
            vertical_similarity.append([words[i], words[j], vec])   
 
    # write excel
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    for idx in range(len(vertical_similarity)):
        ws.append(vertical_similarity[idx])
    wb.save('word_similarity.xlsx')
    
def write_matrix_excel(model):
    # write report (Matrix) 
    co_sin_list = list()
    words = list(model.wv.vocab)
    for i in range(len(words)):
        tmp = list()
        for j in range(len(words)):
            vec = model.wv.similarity(words[i],words[j])
            tmp.append(vec)       
        co_sin_list.append(tmp)
    
    # write report (Matrix)      
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    words.insert(0,' ')
    ws.append(words)
    del words[0]
    for idx in range(len(co_sin_list)):
        tmp_lst = list()
        tmp_lst = co_sin_list[idx]
        tmp_lst.insert(0, words[idx])
        ws.append(tmp_lst)
    wb.save('word_similarity_matrix.xlsx')


def word2vec(tokenized_data):
    
    # Word2Vec parameter
    # Reference https://radimrehurek.com/gensim/models/word2vec.html
    # size: Dimensionality of the word vectors.
    # window: Maximum distance between the current and predicted word within a sentence.
    # min_count: Ignores all words with total frequency lower than this.
    
    model = Word2Vec(tokenized_data, size=100, window=5, min_count=5)
    model.save("word2vec.model")
    #model = Word2Vec.load("word2vec.model")
    
    # Get vocabulary
    words = list(model.wv.vocab)
    
    # Get word similarity
    word_sim = model.wv.similarity('เลี้ยงลูก','มีความสุข')
    print(word_sim)
    
    # Get most word similarity
    most_sim = model.wv.most_similar('ลาออก')
    print(most_sim)
  